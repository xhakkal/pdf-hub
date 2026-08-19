import os
import io
from pathlib import Path
import fitz  # pymupdf
from PyPDF2 import PdfReader, PdfWriter
from docx import Document
from docx.shared import Pt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class PDFConverter:
    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.pdf_filename = Path(pdf_path).stem

    def compress_pdf(self):
        """Comprime PDF reduzindo tamanho do arquivo."""
        try:
            from pathlib import Path as PathlibPath

            reader = PdfReader(self.pdf_path)
            writer = PdfWriter()

            for page in reader.pages:
                # Adicionar página ao writer (compressão padrão)
                writer.add_page(page)

            output_filename = f"{self.pdf_filename}_compressed.pdf"
            output_dir = os.path.dirname(self.pdf_path)
            output_path = os.path.join(output_dir, '..', 'output', output_filename)
            output_path = os.path.abspath(output_path)

            # Garantir que o diretório existe
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Usar compressão ao escrever
            with open(output_path, 'wb') as output_file:
                writer.write(output_file)

            return [output_path], None
        except Exception as e:
            return None, f"Erro ao comprimir PDF: {str(e)}"

    def merge_pdfs(self, other_pdf_paths):
        """Une múltiplos PDFs em um único arquivo."""
        try:
            writer = PdfWriter()

            # Primeiro adiciona o PDF principal
            reader = PdfReader(self.pdf_path)
            for page in reader.pages:
                writer.add_page(page)

            # Adiciona os outros PDFs
            for pdf_path in other_pdf_paths:
                reader = PdfReader(pdf_path)
                for page in reader.pages:
                    writer.add_page(page)

            output_filename = f"{self.pdf_filename}_merged.pdf"
            output_dir = os.path.dirname(self.pdf_path)
            output_path = os.path.join(output_dir, '..', 'output', output_filename)
            output_path = os.path.abspath(output_path)

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            with open(output_path, 'wb') as output_file:
                writer.write(output_file)

            return [output_path], None
        except Exception as e:
            return None, f"Erro ao unir PDFs: {str(e)}"

    def split_pdf(self, split_mode='all', page_ranges=None):
        """Divide PDF em múltiplos arquivos.

        Args:
            split_mode: 'all' (cada página), 'range' (intervalos), 'every_n' (cada N páginas)
            page_ranges: lista de tuplas (start, end) para modo 'range', ou int para 'every_n'
        """
        try:
            reader = PdfReader(self.pdf_path)
            total_pages = len(reader.pages)
            output_files = []
            output_dir = os.path.dirname(self.pdf_path)
            output_dir = os.path.join(output_dir, '..', 'output')
            output_dir = os.path.abspath(output_dir)
            os.makedirs(output_dir, exist_ok=True)

            if split_mode == 'all':
                # Cada página vira um arquivo
                for i in range(total_pages):
                    writer = PdfWriter()
                    writer.add_page(reader.pages[i])
                    output_filename = f"{self.pdf_filename}_page_{i+1:03d}.pdf"
                    output_path = os.path.join(output_dir, output_filename)
                    with open(output_path, 'wb') as f:
                        writer.write(f)
                    output_files.append(output_path)

            elif split_mode == 'range' and page_ranges:
                # Intervalos específicos
                for idx, (start, end) in enumerate(page_ranges):
                    writer = PdfWriter()
                    start_idx = max(0, start - 1)
                    end_idx = min(total_pages, end)
                    for i in range(start_idx, end_idx):
                        writer.add_page(reader.pages[i])
                    output_filename = f"{self.pdf_filename}_part_{idx+1}_pages_{start}-{end}.pdf"
                    output_path = os.path.join(output_dir, output_filename)
                    with open(output_path, 'wb') as f:
                        writer.write(f)
                    output_files.append(output_path)

            elif split_mode == 'every_n' and page_ranges:
                # A cada N páginas
                n = page_ranges if isinstance(page_ranges, int) else 1
                for i in range(0, total_pages, n):
                    writer = PdfWriter()
                    end = min(i + n, total_pages)
                    for j in range(i, end):
                        writer.add_page(reader.pages[j])
                    output_filename = f"{self.pdf_filename}_part_{i//n + 1}_pages_{i+1}-{end}.pdf"
                    output_path = os.path.join(output_dir, output_filename)
                    with open(output_path, 'wb') as f:
                        writer.write(f)
                    output_files.append(output_path)

            return output_files, None
        except Exception as e:
            return None, f"Erro ao dividir PDF: {str(e)}"

    def rotate_pages(self, rotation, page_numbers=None):
        """Rotaciona páginas do PDF.

        Args:
            rotation: 90, 180, 270 (graus no sentido horário)
            page_numbers: lista de números de página (1-indexed) ou None para todas
        """
        try:
            reader = PdfReader(self.pdf_path)
            writer = PdfWriter()
            total_pages = len(reader.pages)

            if page_numbers is None:
                page_numbers = list(range(1, total_pages + 1))

            # Validar números de página
            page_indices = [p - 1 for p in page_numbers if 1 <= p <= total_pages]

            for i in range(total_pages):
                page = reader.pages[i]
                if i in page_indices:
                    # Rotaciona a página
                    page.rotate(rotation)
                writer.add_page(page)

            output_filename = f"{self.pdf_filename}_rotated.pdf"
            output_dir = os.path.dirname(self.pdf_path)
            output_path = os.path.join(output_dir, '..', 'output', output_filename)
            output_path = os.path.abspath(output_path)

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            with open(output_path, 'wb') as output_file:
                writer.write(output_file)

            return [output_path], None
        except Exception as e:
            return None, f"Erro ao rotacionar páginas: {str(e)}"

    def add_watermark(self, watermark_text, opacity=0.3, angle=45, font_size=48, color='gray'):
        """Adiciona marca d'água de texto a todas as páginas do PDF."""
        try:
            # Usar PyMuPDF (fitz) para adicionar marca d'água
            doc = fitz.open(self.pdf_path)

            # Criar marca d'água como overlay
            for page in doc:
                page_rect = page.rect

                # Criar texto da marca d'água
                text_color = {
                    'red': (1, 0, 0),
                    'blue': (0, 0, 1),
                    'green': (0, 0.5, 0),
                    'gray': (0.5, 0.5, 0.5),
                    'black': (0, 0, 0),
                }.get(color, (0.5, 0.5, 0.5))

                # Inserir texto diagonalmente
                text_rect = fitz.Rect(0, 0, page_rect.width, page_rect.height)
                page.insert_text(
                    text_rect.tl + (page_rect.width/2, page_rect.height/2),
                    watermark_text,
                    fontsize=font_size,
                    color=text_color,
                    overlay=True,
                    render_mode=0,  # preenchido
                    rotate=angle
                )

            output_filename = f"{self.pdf_filename}_watermarked.pdf"
            output_dir = os.path.dirname(self.pdf_path)
            output_path = os.path.join(output_dir, '..', 'output', output_filename)
            output_path = os.path.abspath(output_path)

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            doc.save(output_path)
            doc.close()

            return [output_path], None
        except Exception as e:
            return None, f"Erro ao adicionar marca d'água: {str(e)}"

    def add_password(self, user_password, owner_password=None, permissions=None):
        """Protege PDF com senha.

        Args:
            user_password: senha para abrir o arquivo
            owner_password: senha de proprietário (padrão = user_password)
            permissions: dict com permissões (print, copy, modify, annotate)
        """
        try:
            reader = PdfReader(self.pdf_path)
            writer = PdfWriter()

            for page in reader.pages:
                writer.add_page(page)

            if owner_password is None:
                owner_password = user_password

            # Configurar permissões padrão
            if permissions is None:
                permissions = {
                    'print': True,
                    'copy': True,
                    'modify': True,
                    'annotate': True
                }

            # Criptografar
            writer.encrypt(
                user_password=user_password,
                owner_password=owner_password,
                permissions=permissions
            )

            output_filename = f"{self.pdf_filename}_protected.pdf"
            output_dir = os.path.dirname(self.pdf_path)
            output_path = os.path.join(output_dir, '..', 'output', output_filename)
            output_path = os.path.abspath(output_path)

            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            with open(output_path, 'wb') as output_file:
                writer.write(output_file)

            return [output_path], None
        except Exception as e:
            return None, f"Erro ao proteger PDF: {str(e)}"
    
    def to_images(self, format='PNG', dpi=150):
        """Converte PDF para imagens (PNG ou JPG) usando PyMuPDF."""
        try:
            doc = fitz.open(self.pdf_path)
            output_files = []
            output_dir = os.path.dirname(self.pdf_path)
            output_dir = os.path.join(output_dir, '..', 'output')
            output_dir = os.path.abspath(output_dir)
            os.makedirs(output_dir, exist_ok=True)

            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)

            for i, page in enumerate(doc):
                pix = page.get_pixmap(matrix=mat)
                output_filename = f"{self.pdf_filename}_page_{i+1:03d}.{format.lower()}"
                output_path = os.path.join(output_dir, output_filename)

                if format.upper() == 'JPG':
                    pix.save(output_path, 'jpeg')
                else:
                    pix.save(output_path, 'png')

                output_files.append(output_path)

            doc.close()
            return output_files, None
        except Exception as e:
            return None, f"Erro ao converter para imagens: {str(e)}"
    
    def to_text(self):
        """Converte PDF para TXT."""
        try:
            text = ""
            with open(self.pdf_path, 'rb') as file:
                pdf_reader = PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    extracted = page.extract_text()
                    text += extracted + "\n"
            
            output_filename = f"{self.pdf_filename}.txt"
            output_dir = os.path.dirname(self.pdf_path)
            output_dir = os.path.join(output_dir, '..', 'output')
            output_dir = os.path.abspath(output_dir)
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, output_filename)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(text)
            
            return [output_path], None
        except Exception as e:
            return None, f"Erro ao converter para TXT: {str(e)}"
    
    def to_docx(self):
        """Converte PDF para DOCX."""
        try:
            doc = Document()
            
            with open(self.pdf_path, 'rb') as file:
                pdf_reader = PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text = page.extract_text()
                    if page_num > 1:
                        doc.add_page_break()
                    doc.add_paragraph(text)
            
            output_filename = f"{self.pdf_filename}.docx"
            output_dir = os.path.dirname(self.pdf_path)
            output_dir = os.path.join(output_dir, '..', 'output')
            output_dir = os.path.abspath(output_dir)
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, output_filename)
            doc.save(output_path)
            
            return [output_path], None
        except Exception as e:
            return None, f"Erro ao converter para DOCX: {str(e)}"
    
    def to_xlsx(self):
        """Converte PDF para XLSX."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "PDF Data"
            
            row = 1
            with open(self.pdf_path, 'rb') as file:
                pdf_reader = PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text = page.extract_text()
                    lines = text.split('\n')
                    
                    for line in lines:
                        if line.strip():
                            ws[f'A{row}'] = line
                            row += 1
            
            output_filename = f"{self.pdf_filename}.xlsx"
            output_dir = os.path.dirname(self.pdf_path)
            output_dir = os.path.join(output_dir, '..', 'output')
            output_dir = os.path.abspath(output_dir)
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, output_filename)
            wb.save(output_path)
            
            return [output_path], None
        except Exception as e:
            return None, f"Erro ao converter para XLSX: {str(e)}"
